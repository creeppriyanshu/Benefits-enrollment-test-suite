"""Generates test_cases.xlsx — the input test case sheet for the test runner."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = [
    "Test_ID", "Description", "Endpoint", "Method", "Input_JSON",
    "Expected_Status", "Expected_Contains",
    "Actual_Status", "Actual_Response", "Result"
]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF")
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

rows = [
    ["TC01", "Eligible active employee passes eligibility check",
     "/eligibility/check", "POST", {"employee_id": "E001"}, 200, "true"],

    ["TC02", "Employee still in 30-day waiting period is rejected",
     "/eligibility/check", "POST", {"employee_id": "E002"}, 200, "waiting period"],

    ["TC03", "Terminated employee fails eligibility check",
     "/eligibility/check", "POST", {"employee_id": "E003"}, 200, "not in Active"],

    ["TC04", "Eligibility check for unknown employee returns 404",
     "/eligibility/check", "POST", {"employee_id": "E999"}, 404, "not found"],

    ["TC05", "Spouse dependent is eligible",
     "/dependents/E001", "GET", {}, 200, "Spouse is eligible"],

    ["TC06", "Child dependent under 26 is eligible",
     "/dependents/E001", "GET", {}, 200, "under age 26"],

    ["TC07", "Child dependent aged 27 has aged out",
     "/dependents/E004", "GET", {}, 200, "aged out"],

    ["TC08", "Dependents lookup for unknown employee returns 404",
     "/dependents/E999", "GET", {}, 404, "not found"],

    ["TC09", "Enrollment succeeds inside open enrollment window",
     "/enrollment/submit", "POST",
     {"employee_id": "E001", "plan_id": "PLAN-GOLD", "enrollment_date": "2026-11-05"},
     200, "Enrolled"],

    ["TC10", "Enrollment outside window without QLE is rejected",
     "/enrollment/submit", "POST",
     {"employee_id": "E001", "plan_id": "PLAN-GOLD", "enrollment_date": "2026-06-01"},
     200, "outside the open enrollment window"],

    ["TC11", "Enrollment outside window WITH valid QLE is accepted",
     "/enrollment/submit", "POST",
     {"employee_id": "E001", "plan_id": "PLAN-GOLD", "enrollment_date": "2026-06-01", "is_qle": True},
     200, "Enrolled"],

    ["TC12", "Enrollment rejected for employee who fails eligibility",
     "/enrollment/submit", "POST",
     {"employee_id": "E002", "plan_id": "PLAN-GOLD", "enrollment_date": "2026-11-05"},
     200, "Rejected"],

    ["TC13", "Enrollment missing plan_id returns 400",
     "/enrollment/submit", "POST",
     {"employee_id": "E001", "enrollment_date": "2026-11-05"},
     400, "required"],

    ["TC14", "Valid life event returns a 30-day SEP window",
     "/life-event", "POST",
     {"employee_id": "E001", "event_type": "birth", "event_date": "2026-07-01"},
     200, "2026-07-31"],

    ["TC15", "Invalid life event type returns 400",
     "/life-event", "POST",
     {"employee_id": "E001", "event_type": "promotion", "event_date": "2026-07-01"},
     400, "Invalid event_type"],

    ["TC16", "COBRA eligible for terminated employee with prior enrollment",
     "/cobra/check", "POST",
     {"employee_id": "E003", "termination_date": "2026-05-15"},
     200, "Eligible for COBRA"],

    ["TC17", "COBRA not eligible for employer with under 20 employees",
     "/cobra/check", "POST",
     {"employee_id": "E004", "termination_date": "2026-06-01"},
     200, "fewer than 20 employees"],

    ["TC18", "COBRA check missing termination_date returns 400",
     "/cobra/check", "POST",
     {"employee_id": "E003"},
     400, "required"],
]

for row in rows:
    test_id, desc, endpoint, method, input_json, exp_status, exp_contains = row
    ws.append([test_id, desc, endpoint, method, json.dumps(input_json), exp_status, exp_contains, "", "", ""])

# Column widths
widths = {"A": 10, "B": 46, "C": 20, "D": 8, "E": 42, "F": 14, "G": 30, "H": 14, "I": 50, "J": 10}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# Wrap text for long columns, apply Arial font to all data rows
wrap_cols = {"B", "E", "I"}
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name="Arial", size=10)
        if col_letter in wrap_cols:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.alignment = Alignment(vertical="top")

ws.freeze_panes = "A2"

wb.save("test_cases.xlsx")
print("test_cases.xlsx created with", ws.max_row - 1, "test cases")
