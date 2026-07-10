"""
Benefits Enrollment Test Runner
--------------------------------
Reads test cases from test_cases.xlsx, executes them against the mock
H&W Enrollment API (app.py, must be running on http://127.0.0.1:5000),
writes actual results back into the workbook, and generates an HTML
pass/fail report (report.html).

Run:
    1) python app.py                 (in one terminal)
    2) python test_runner.py         (in another terminal)
"""

import json
import requests
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Template

BASE_URL = "http://127.0.0.1:5000"
INPUT_FILE = "test_cases.xlsx"
REPORT_FILE = "report.html"

COLS = {
    "test_id": 1, "description": 2, "endpoint": 3, "method": 4, "input_json": 5,
    "expected_status": 6, "expected_contains": 7,
    "actual_status": 8, "actual_response": 9, "result": 10,
}


def run_tests():
    wb = load_workbook(INPUT_FILE)
    ws = wb["Test Cases"]

    results = []
    passed = failed = 0

    for row in range(2, ws.max_row + 1):
        test_id = ws.cell(row=row, column=COLS["test_id"]).value
        if not test_id:
            continue

        description = ws.cell(row=row, column=COLS["description"]).value
        endpoint = ws.cell(row=row, column=COLS["endpoint"]).value
        method = ws.cell(row=row, column=COLS["method"]).value
        input_json_raw = ws.cell(row=row, column=COLS["input_json"]).value
        expected_status = ws.cell(row=row, column=COLS["expected_status"]).value
        expected_contains = str(ws.cell(row=row, column=COLS["expected_contains"]).value or "")

        try:
            payload = json.loads(input_json_raw) if input_json_raw else {}
        except json.JSONDecodeError:
            payload = {}

        url = BASE_URL + endpoint
        try:
            if method.upper() == "GET":
                resp = requests.get(url, timeout=5)
            else:
                resp = requests.post(url, json=payload, timeout=5)
            actual_status = resp.status_code
            actual_body = resp.text
        except requests.exceptions.ConnectionError:
            actual_status = "ERROR"
            actual_body = "Could not connect to API. Is app.py running on port 5000?"

        status_ok = (actual_status == expected_status)
        contains_ok = expected_contains.lower() in actual_body.lower()
        test_passed = status_ok and contains_ok
        result_text = "PASS" if test_passed else "FAIL"

        if test_passed:
            passed += 1
        else:
            failed += 1

        # write back into the sheet
        ws.cell(row=row, column=COLS["actual_status"]).value = actual_status
        ws.cell(row=row, column=COLS["actual_response"]).value = actual_body[:500]
        ws.cell(row=row, column=COLS["result"]).value = result_text

        results.append({
            "test_id": test_id,
            "description": description,
            "endpoint": endpoint,
            "method": method,
            "expected_status": expected_status,
            "actual_status": actual_status,
            "result": result_text,
        })

    wb.save(INPUT_FILE)
    return results, passed, failed


REPORT_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Benefits Enrollment - Test Execution Report</title>
<style>
  body { font-family: Arial, sans-serif; margin: 40px; color: #1a1a1a; background: #f7f8fa; }
  h1 { margin-bottom: 4px; }
  .meta { color: #555; margin-bottom: 24px; }
  .summary { display: flex; gap: 16px; margin-bottom: 28px; }
  .card { padding: 16px 24px; border-radius: 8px; color: white; min-width: 120px; text-align: center; }
  .total { background: #1F4E78; }
  .pass { background: #2e7d32; }
  .fail { background: #c62828; }
  .card .num { font-size: 28px; font-weight: bold; display: block; }
  table { width: 100%; border-collapse: collapse; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
  th, td { text-align: left; padding: 10px 12px; border-bottom: 1px solid #e0e0e0; font-size: 14px; }
  th { background: #1F4E78; color: white; }
  tr:hover { background: #f1f5fa; }
  .pass-cell { color: #2e7d32; font-weight: bold; }
  .fail-cell { color: #c62828; font-weight: bold; }
</style>
</head>
<body>
  <h1>Benefits Enrollment (H&amp;W) - Test Execution Report</h1>
  <div class="meta">Generated {{ timestamp }}</div>

  <div class="summary">
    <div class="card total"><span class="num">{{ total }}</span>Total</div>
    <div class="card pass"><span class="num">{{ passed }}</span>Passed</div>
    <div class="card fail"><span class="num">{{ failed }}</span>Failed</div>
  </div>

  <table>
    <tr>
      <th>Test ID</th><th>Description</th><th>Endpoint</th><th>Method</th>
      <th>Expected Status</th><th>Actual Status</th><th>Result</th>
    </tr>
    {% for r in results %}
    <tr>
      <td>{{ r.test_id }}</td>
      <td>{{ r.description }}</td>
      <td>{{ r.endpoint }}</td>
      <td>{{ r.method }}</td>
      <td>{{ r.expected_status }}</td>
      <td>{{ r.actual_status }}</td>
      <td class="{{ 'pass-cell' if r.result == 'PASS' else 'fail-cell' }}">{{ r.result }}</td>
    </tr>
    {% endfor %}
  </table>
</body>
</html>
"""


def generate_report(results, passed, failed):
    template = Template(REPORT_TEMPLATE)
    html = template.render(
        results=results,
        total=passed + failed,
        passed=passed,
        failed=failed,
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    with open(REPORT_FILE, "w") as f:
        f.write(html)


if __name__ == "__main__":
    print("Running test cases against", BASE_URL, "...")
    results, passed, failed = run_tests()
    generate_report(results, passed, failed)
    print(f"Done. {passed} passed, {failed} failed out of {passed + failed}.")
    print(f"Report written to {REPORT_FILE}")
    print(f"Results written back into {INPUT_FILE}")
